#!/usr/bin/env python3

import os
import openpyxl

# Rutas de directorios
simulation_results_dir = os.path.expanduser("~/mySimTools/gem5/resources_assignment/simulation_results")
excel_output_path = os.path.expanduser("~/mySimTools/gem5/resources_assignment/IPC_results.xlsx")

# Recopilación de datos
data_by_application = {}

# Recorrer todos los archivos en el directorio
for file_name in os.listdir(simulation_results_dir):
    # Verificar que el archivo sea .txt
    if file_name.endswith(".txt"):
        # Extraer la configuración del archivo
        parts = file_name.split('_')
        try:
            sim_num = parts[1]
            app_type = parts[2]  # tipo de aplicación (jpg2k, h264, mp3, etc.)
            config_part = "_".join(parts[3:]).split('.')[0]  # el resto de la configuración
            
            # Filtrar configuraciones que coincidan con nuestras combinaciones
            key = f"{app_type}_{parts[3]}"  # Usa la parte que indica si es enc o dec
            if key in data_by_application:
                data_by_application[key].append((config_part, file_name))
            else:
                data_by_application[key] = [(config_part, file_name)]

        except IndexError:
            print(f"Error al procesar el archivo {file_name}, formato inesperado.")
            continue

# Crear o cargar el archivo Excel
if os.path.exists(excel_output_path):
    workbook = openpyxl.load_workbook(excel_output_path)
else:
    workbook = openpyxl.Workbook()

# Procesar cada aplicación y modo
for app_type, config_files in data_by_application.items():
    # Crear una hoja nueva para cada aplicación y modo si no existe
    if app_type not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=app_type)
        sheet.append(["Configuración", "Archivo TXT", "IPC"])
    else:
        sheet = workbook[app_type]

    # Extraer valores de IPC y escribir en la hoja
    for config_part, file_name in config_files:
        file_path = os.path.join(simulation_results_dir, file_name)

        try:
            # Leer la línea 18 y extraer el valor de IPC
            with open(file_path, 'r') as f:
                lines = f.readlines()
                ipc_line = lines[17]  # línea 18 es el índice 17
                ipc_value = float(ipc_line.split()[1])  # asume que el valor de IPC es el segundo elemento

                # Guardar datos en la hoja
                sheet.append([config_part, file_name, ipc_value])
                print(f"Datos guardados para {file_name} en la configuración {app_type}")

        except (IndexError, ValueError, FileNotFoundError) as e:
            print(f"Error al leer el archivo {file_name}: {e}")
            continue

# Guardar el archivo Excel
workbook.save(excel_output_path)
print(f"Todos los datos de IPC se han guardado en {excel_output_path}")
