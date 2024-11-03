#!/usr/bin/env python3

import os
import openpyxl

# Ruta del directorio donde están los resultados de simulación
mcpat_results_dir = os.path.expanduser("~/mySimTools/gem5/resources_assignment/simulation_results")

# Archivo Excel donde se guardarán los tiempos
excel_file_path = os.path.expanduser("~/mySimTools/gem5/resources_assignment/arm_execution_times.xlsx")

# Crear o cargar el archivo Excel
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
else:
    workbook = openpyxl.Workbook()

# Seleccionar o crear una hoja de trabajo
sheet = workbook.active
sheet.title = "Tiempos de Ejecución"
if sheet.max_row == 1:  # Añadir encabezados si la hoja está vacía
    sheet.append(["Simulación", "Archivo TXT", "Tiempo de Ejecución (segundos)"])

# Procesar cada archivo TXT desde sim1 hasta sim288
for sim_num in range(1, 289):
    # Buscar el archivo TXT correspondiente que comienza con "stats_sim" seguido del número
    txt_file = next(
        (os.path.join(mcpat_results_dir, file_name)
         for file_name in os.listdir(mcpat_results_dir)
         if file_name.startswith(f"stats_sim{sim_num}_") and file_name.endswith(".txt")),
        None
    )

    # Verificar si el archivo TXT existe
    if txt_file:
        # Extraer solo el nombre del archivo sin extensión
        txt_file_name = os.path.basename(txt_file).replace(".txt", "")
        
        try:
            # Abrir el archivo y leer la línea 7
            with open(txt_file, 'r') as file:
                lines = file.readlines()
                # Extraer el valor numérico de hostSeconds en la línea 7
                host_seconds = float(lines[16].split()[1])

            # Guardar el tiempo de ejecución en el archivo Excel
            sheet.append([f"sim{sim_num}", txt_file_name, host_seconds])
            print(f"Tiempo de ejecución para sim{sim_num}: {host_seconds:.2f} ICPs\n")

        except (IndexError, ValueError) as e:
            print(f"Error al leer hostSeconds para sim{sim_num}: {e}")
    else:
        print(f"Archivo TXT para sim{sim_num} no encontrado.")

# Guardar y cerrar el archivo Excel
workbook.save(excel_file_path)
print(f"Todos los tiempos de ejecución se han guardado en {excel_file_path}")
