#!/usr/bin/env python3

import os
import openpyxl

# Rutas de directorios
simulation_results_dir = os.path.expanduser("~/mySimTools/gem5/resources_assignment/simulation_results")
excel_output_path = os.path.expanduser("~/mySimTools/gem5/resources_assignment/IPC_results.xlsx")

# Configuraciones posibles
L1d_sizes = [16, 64, 256, 512]
write_sizes = [64, 32, 16]
L2_associativities = [4, 8, 2, 16]

# Recopilación de datos
data_by_config = {}

# Recorrer todos los archivos en el directorio
for file_name in os.listdir(simulation_results_dir):
    # Verificar que el archivo sea .txt
    if file_name.endswith(".txt"):
        # Extraer la configuración del archivo
        parts = file_name.split('_')
        try:
            sim_num = parts[1]
            config_part = "_".join(parts[2:]).split('.')[0]
            
            # Filtrar configuraciones que coincidan con nuestras combinaciones
            for L1d in L1d_sizes:
                for write in write_sizes:
                    for L2 in L2_associativities:
                        config_str = f"L1d{L1d}kB_write{write}_L2asso{L2}"
                        
                        if config_str in config_part:
                            config_files = data_by_config.setdefault(config_str, [])
                            config_files.append(file_name)
                            break  # Si encontramos una coincidencia, no es necesario seguir buscando

        except IndexError:
            print(f"Error al procesar el archivo {file_name}, formato inesperado.")
            continue

# Crear o cargar el archivo Excel
if os.path.exists(excel_output_path):
    workbook = openpyxl.load_workbook(excel_output_path)
else:
    workbook = openpyxl.Workbook()

# Procesar cada configuración y extraer datos
for config_str, files in data_by_config.items():
    # Crear una hoja nueva para cada configuración
    if config_str not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=config_str)
        sheet.append(["Simulación", "Archivo TXT", "IPC"])
    else:
        sheet = workbook[config_str]

    # Extraer valores de IPC y escribir en la hoja
    for file_name in files:
        file_path = os.path.join(simulation_results_dir, file_name)
        
        try:
            # Leer la línea 18 y extraer el valor de IPC
            with open(file_path, 'r') as f:
                lines = f.readlines()
                ipc_line = lines[17]  # línea 18 es el índice 17
                ipc_value = float(ipc_line.split()[1])  # asume que el valor de IPC es el segundo elemento

                # Extraer nombre de simulación
                sim_label = file_name.split('_', 2)[2].split('.')[0]  # Parte posterior del nombre
                
                # Guardar en la hoja
                sheet.append([f"sim{sim_num}", file_name, ipc_value])
                print(f"Datos guardados para {file_name} en la configuración {config_str}")

        except (IndexError, ValueError, FileNotFoundError) as e:
            print(f"Error al leer el archivo {file_name}: {e}")
            continue

# Guardar el archivo Excel
workbook.save(excel_output_path)
print(f"Todos los datos de IPC se han guardado en {excel_output_path}")
