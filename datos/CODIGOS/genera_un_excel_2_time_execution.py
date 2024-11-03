#!/usr/bin/env python3

import os
from openpyxl import Workbook, load_workbook

# Rutas de directorios
simulation_results_dir = os.path.expanduser("~/mySimTools/gem5/resources_assignment/simulation_results")
excel_output_path = os.path.expanduser("~/mySimTools/gem5/resources_assignment/ExecutionTime_results.xlsx")

# Recopilación de datos
data_by_application = {}

# Recorrer todos los archivos en el directorio
for file_name in os.listdir(simulation_results_dir):
    # Verificar que el archivo sea .txt
    if file_name.endswith(".txt"):
        # Extraer la configuración del archivo
        parts = file_name.split('_')
        try:
            sim_num = parts[1]
            app_type = parts[2]  # tipo de aplicación (jpg2k, h264, mp3, etc.)
            config_part = "_".join(parts[3:]).split('.')[0]  # el resto de la configuración
            
            # Filtrar configuraciones que coincidan con nuestras combinaciones
            key = f"{app_type}_{parts[3]}"  # Usa la parte que indica si es enc o dec
            if key in data_by_application:
                data_by_application[key].append((config_part, file_name))
            else:
                data_by_application[key] = [(config_part, file_name)]

        except IndexError:
            print(f"Error al procesar el archivo {file_name}, formato inesperado.")
            continue

# Crear o cargar el archivo Excel
if os.path.exists(excel_output_path):
    workbook = load_workbook(excel_output_path)
else:
    workbook = Workbook()

# Procesar cada aplicación y modo
for app_type, config_files in data_by_application.items():
    # Crear una hoja nueva para cada aplicación y modo si no existe
    if app_type not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=app_type)
        sheet.append(["Configuración", "Archivo TXT", "Tiempo de Ejecución (segundos)"])
    else:
        sheet = workbook[app_type]

    # Extraer tiempos de ejecución y escribir en la hoja
    for config_part, file_name in config_files:
        file_path = os.path.join(simulation_results_dir, file_name)

        try:
            # Leer la línea 3 y extraer el valor de tiempo de ejecución (simSeconds)
            with open(file_path, 'r') as f:
                lines = f.readlines()
                exec_time_line = lines[2]  # línea 3 es el índice 2
                exec_time_value = float(exec_time_line.split()[1])  # asume que el valor de simSeconds es el segundo elemento

                # Guardar datos en la hoja
                sheet.append([config_part, file_name, exec_time_value])
                print(f"Datos guardados para {file_name} en la configuración {app_type}")

        except (IndexError, ValueError, FileNotFoundError) as e:
            print(f"Error al leer el archivo {file_name}: {e}")
            continue

# Guardar el archivo Excel
workbook.save(excel_output_path)
print(f"Todos los datos de tiempo de ejecución se han guardado en {excel_output_path}")
