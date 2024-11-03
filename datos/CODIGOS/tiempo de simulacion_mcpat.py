#!/usr/bin/env python3

import os
import subprocess
import time
import openpyxl

# Rutas de directorios
mcpat_executable = os.path.expanduser("~/mySimTools/mcpat/mcpat")
mcpat_results_dir = os.path.expanduser("~/mySimTools/gem5/resources_assignment/McPAT_results")

# Archivo Excel donde se guardarán los tiempos
excel_file_path = os.path.expanduser("~/mySimTools/gem5/resources_assignment/mcpat_execution_times.xlsx")

# Crear o cargar el archivo Excel
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
else:
    workbook = openpyxl.Workbook()

# Seleccionar o crear una hoja de trabajo
sheet = workbook.active
sheet.title = "Tiempos de Ejecución"
if sheet.max_row == 1:  # Añadir encabezados si la hoja está vacía
    sheet.append(["Simulación", "Archivo XML", "Tiempo de Ejecución (segundos)"])

# Ejecutar el comando mcpat para cada archivo XML desde sim1 hasta sim288
for sim_num in range(1, 289):
    # Buscar el archivo XML correspondiente que comienza con "config_sim" seguido del número
    xml_file = next(
        (os.path.join(mcpat_results_dir, file_name)
         for file_name in os.listdir(mcpat_results_dir)
         if file_name.startswith(f"config_sim{sim_num}_") and file_name.endswith(".xml")),
        None
    )

    # Verificar si el archivo XML existe antes de ejecutar el comando
    if xml_file:
        # Extraer solo el nombre del archivo sin extensión
        xml_file_name = os.path.basename(xml_file).replace(".xml", "")

        # Construir el comando
        command = [mcpat_executable, "-infile", xml_file]

        try:
            print(f"Ejecutando mcpat para sim{sim_num} con archivo XML: {xml_file}\n")
            
            # Tomar el tiempo de inicio
            start_time = time.time()
            
            # Ejecutar el comando y esperar a que termine
            subprocess.run(command, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            
            # Calcular el tiempo de ejecución
            elapsed_time = time.time() - start_time

            # Guardar el tiempo de ejecución en el archivo Excel
            sheet.append([f"sim{sim_num}", xml_file_name, elapsed_time])
            print(f"Tiempo de ejecución para sim{sim_num}: {elapsed_time:.2f} segundos\n")

        except subprocess.CalledProcessError as e:
            print(f"Error al ejecutar mcpat para sim{sim_num}: {e}")
    else:
        print(f"Archivo XML para sim{sim_num} no encontrado.")

# Guardar y cerrar el archivo Excel
workbook.save(excel_file_path)
print(f"Todos los tiempos de ejecución se han guardado en {excel_file_path}")
