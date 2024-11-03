#!/usr/bin/env python3

import os
import openpyxl
import json

# Ruta del directorio donde están los resultados de simulación
mcpat_results_dir = os.path.expanduser("~/mySimTools/McPAT_json_results")

# Archivo Excel donde se guardarán los tiempos y variables
excel_file_path = os.path.expanduser("~/mySimTools/json_variables.xlsx")

# Crear o cargar el archivo Excel
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
else:
    workbook = openpyxl.Workbook()

# Seleccionar o crear una hoja de trabajo
sheet = workbook.active
sheet.title = "Variables de Simulación"
if sheet.max_row == 1:  # Añadir encabezados si la hoja está vacía
    sheet.append(["Simulación", "Archivo JSON", "Total Leakage", "Runtime Dynamic"])

# Procesar cada archivo JSON desde sim1 hasta sim288
for sim_num in range(1, 289):
    # Buscar el archivo JSON correspondiente que comienza con "config_sim" seguido del número
    json_file = next(
        (os.path.join(mcpat_results_dir, file_name)
         for file_name in os.listdir(mcpat_results_dir)
         if file_name.startswith(f"config_sim{sim_num}_") and file_name.endswith(".json")),
        None
    )

    # Verificar si el archivo JSON existe
    if json_file:
        # Extraer solo el nombre del archivo sin extensión
        json_file_name = os.path.basename(json_file).replace(".json", "")
        
        try:
            # Abrir el archivo JSON y leer las variables requeridas
            with open(json_file, 'r') as file:
                data = json.load(file)
                # Obtener los valores de "Total Leakage" y "Runtime Dynamic"
                total_leakage = data.get("Total Leakage", None)
                runtime_dynamic = data.get("Runtime Dynamic", None)

            # Guardar los resultados en el archivo Excel
            sheet.append([f"sim{sim_num}", json_file_name, total_leakage, runtime_dynamic])
            print(f"Variables para sim{sim_num}: Total Leakage = {total_leakage}, Runtime Dynamic = {runtime_dynamic}\n")

        except (json.JSONDecodeError, ValueError) as e:
            print(f"Error al leer las variables para sim{sim_num}: {e}")
    else:
        print(f"Archivo JSON para sim{sim_num} no encontrado.")

# Guardar y cerrar el archivo Excel
workbook.save(excel_file_path)
print(f"Todos los datos se han guardado en {excel_file_path}")
