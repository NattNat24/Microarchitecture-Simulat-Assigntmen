#!/usr/bin/env python3

import os
import openpyxl

# Ruta del directorio donde están los resultados de simulación
mcpat_results_dir = os.path.expanduser("~/mySimTools/simulation_results")

# Archivo Excel donde se guardarán los tiempos
excel_file_path = os.path.expanduser("~/mySimTools/stats_cpi.xlsx")

# Crear o cargar el archivo Excel
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
else:
    workbook = openpyxl.Workbook()

# Seleccionar o crear una nueva hoja de trabajo para CPI
if "CPI Valores" in workbook.sheetnames:
    sheet = workbook["CPI Valores"]
    # Limpiar la hoja existente
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=sheet.max_row):
        for cell in row:
            cell.value = None
else:
    sheet = workbook.create_sheet("CPI Valores")

# Añadir encabezados si la hoja está vacía
if sheet.max_row == 1:  
    sheet.append(["Simulación", "Archivo TXT", "CPI"])

# Procesar cada archivo TXT desde sim1 hasta sim288
for sim_num in range(1, 289):
    # Buscar el archivo TXT correspondiente que comienza con "stats_sim" seguido del número
    txt_file = next(
        (os.path.join(mcpat_results_dir, file_name)
         for file_name in os.listdir(mcpat_results_dir)
         if file_name.startswith(f"stats_sim{sim_num}_") and file_name.endswith(".txt")),
        None
    )

    # Verificar si el archivo TXT existe
    if txt_file:
        # Extraer solo el nombre del archivo sin extensión
        txt_file_name = os.path.basename(txt_file).replace(".txt", "")
        
        try:
            # Abrir el archivo y buscar la línea que contiene el CPI
            with open(txt_file, 'r') as file:
                lines = file.readlines()
                # Buscar la línea que contiene "CPI"
                cpi_line = next((line for line in lines if "CPI" in line), None)
                
                if cpi_line:
                    # Extraer el valor numérico de CPI
                    cpi_value = float(cpi_line.split()[1])  # Asumiendo que el valor está en la segunda posición
                    
                    # Guardar el valor de CPI en el archivo Excel
                    sheet.append([f"sim{sim_num}", txt_file_name, cpi_value])
                    print(f"CPI para sim{sim_num}: {cpi_value:.2f}\n")
                else:
                    print(f"CPI no encontrado en el archivo {txt_file_name}.")

        except (IndexError, ValueError) as e:
            print(f"Error al leer CPI para sim{sim_num}: {e}")
    else:
        print(f"Archivo TXT para sim{sim_num} no encontrado.")

# Guardar y cerrar el archivo Excel
workbook.save(excel_file_path)
print(f"Todos los valores de CPI se han guardado en {excel_file_path}")
